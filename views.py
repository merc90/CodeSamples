import openpyxl
import zipfile
import io
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import Alignment
from cw.eee_app_ini import *


class T4IndexView(View):
    '''
        Index page, displays the list of supervisees
    '''
    
    def get(self, request):
        user = CommonFunctions.getUserLogin(request, T4Functions.suAccess)["user"]
        CID = CommonFunctions.getCID(user)
        data = {}
        eeID = str(CommonFunctions.getEEID(user))
        data["authorized"] = T4Functions.isAuthorized(eeID)
        data.update({"breadcrumb": '<li><a>Home</a></li>',
        "pageTitle": "T4Monitoring Home"})
        data.update(CommonFunctions.getUserLogin(request, T4Functions.suAccess))
        data["empString"] = "0000-00"
        if data["authorized"]:
            with connection.cursor() as cursor:
                query = "SELECT eeid, CID, login, Namei, FeeClass, ERegyr, Alias FROM \
                    eedbo.eepx WHERE (EReg = 'P' OR EReg = 'F'  OR EReg = 'X') and (PSup1 = " + eeID + " or \
                    PSup2 = " + eeID + ") and SUBSTRING(ERegyr, 2, 2) < 6"
                data["pgrList"] = cursor.execute(query).fetchall()
                query = "SELECT eeid, CID, login, Namei, FeeClass, ERegyr, \
                    Alias, A.* FROM TutorAllocation A LEFT JOIN eedbo.eepx B \
                    ON A. tutor = b.eeid where tutee = " + eeID
                data["tuteeList"] = cursor.execute(query).fetchall()
                data["user"] = CommonFunctions.getUser(user)
        elif user in T4Functions.suAccess:
            return redirect("/t4monitoring/report"+data["urlEncode"])
        template = loader.render_to_string('app/index.html', {"data": data})
        return HttpResponse(template)

    def post(self, request):
        return HttpResponse("POST")
        
class T4GetDetailsAPIView(View):
    '''
        API, return the log information for a required student for the required month
        
        Parameters:
            login (string): username
            period (string): month in the format YYYY-MM
    '''
    
    def get(self, request, login: str, period: str):
        return JsonResponse({"info": T4Functions.getPeriodInfo(login, period)})

    def post(self, request, login: str, period: str):
        return HttpResponse("POST")

class T4ContactView(View):
    '''
        Form page, display the information for the required student for the 
        required month if information exist otherwise new information can be
        entered.
        
        Parameters:
            pgrID (int): EEID
            period (string): month in the format YYYY-MM
    '''
        
    def get(self, request, pgrID: int, period: str):
        user = CommonFunctions.getUserLogin(request, T4Functions.suAccess)["user"]
        CID = CommonFunctions.getCID(user)
        eeID = str(CommonFunctions.getEEID(user))
        data = {}
        data.update(CommonFunctions.getUserLogin(request, T4Functions.suAccess))
        data["pageTitle"] = "T4 Contacts"
        data["breadcrumb"] = '<li><a href="/t4monitoring/' + data["urlEncode"]\
            + '"> Home</a></li><li><a>' + data["pageTitle"] + '</a></li>'
        data["authorized"] = T4Functions.isAuthorized(eeID)
        if data["authorized"]:
            with connection.cursor() as cursor:
                query = "SELECT eeid, CID, login, Namei, FeeClass, ERegyr, formal FROM \
                    eedbo.eepx WHERE eeid = " + str(pgrID) + " and (PSup1 = " + \
                    eeID + " or PSup2 = " + eeID + ")"
                data["pgrCand"] = cursor.execute(query).fetchone()
                if not data["pgrCand"]:
                    return redirect("/t4monitoring/"+data["urlEncode"])
                else:
                    data["formal"] = CommonFunctions.getFormal(user)
                    data["user"] = CommonFunctions.getUser(user)
                    data.update(T4Functions.getLatestPeriod(data["pgrCand"][2]))
                data["pgrCand"][1] = "{:08d}".format(data["pgrCand"][1])
                cursor.close()
            if period != "0000-00":
                data["year"] = period[0:4]
                data["month"] = period[5:7]
            data["info"] = T4Functions.getPeriodInfo(data["pgrCand"][2], str(data["year"])+"-"+data["month"])
            data.update(T4Functions.getIncompleteList(data["pgrCand"][2]))
            template = loader.render_to_string('app/contacts.html', {"data": data})
            return HttpResponse(template)
        else:
            return redirect("/t4monitoring/"+data["urlEncode"])

    def post(self, request, pgrID:int, period: str):
        with connection.cursor() as cursor:
            data = {}
            maps = { 1 : (1,1), 2: (1,2), 3 : (2,1), 4 : (2,2), 5 : (3,0)}
            data.update(CommonFunctions.getUserLogin(request, T4Functions.suAccess))
            period = request.POST.getlist("period")[0]
            checks = request.POST.getlist("checks[]")
            reason = request.POST.getlist("reason")
            reason = reason[0] if reason else ""
            period = period[3:] + "-" + period[:2]
            studentLogin = CommonFunctions.getLogin(pgrID)
            updatedBy = CommonFunctions.getUserLogin(request, T4Functions.suAccess)["user"]
            query = "DELETE FROM ElecEngwww.tier4log WHERE studentLogin = '" \
                + studentLogin + "' and logPeriod = '" + period + "'"
            cursor.execute(query)
            for check in checks:
                (contactType, count) = maps.get(check, (4,0))
                query = "INSERT INTO ElecEngwww.tier4log (studentLogin, \
                    logPeriod, contactType, contactCount, updateBy, note) VALUES \
                    ('" + studentLogin + "', '" + period + "', " + str(contactType) + ", \
                    " + str(count) + ", '" + updatedBy + "', '" + reason + "')"
                cursor.execute(query)
            cursor.close()
        return redirect("/t4monitoring/contacts/"+pgrID+"/0000-00"+data["urlEncode"])
        
class T4Functions():
    '''
        App specific functions
    '''
    
    '''
    import list of superusers from eee_app_ini file
    '''
    suAccess = superUser["t4monitoring"]
    
    months = {"Jan":"01", "Feb":"02", "Mar":"03", "Apr":"04", "May":"05", \
            "Jun":"06", "Jul":"07", "Aug":"08", "Sep":"09", "Oct":"10", \
            "Nov":"11", "Dec":"12"}
    monthsRev = {"01":"Jan", "02":"Feb", "03":"Mar", "04":"Apr", "05":"May", \
        "06":"Jun", "07":"Jul", "08":"Aug", "09":"Sep", "10":"Oct", \
        "11":"Nov", "12":"Dec"}

    year = datetime.datetime.today().year - 1 \
        if datetime.datetime.today().month < 10 else \
        datetime.datetime.today().year
        
    def getPeriodInfo(login: str, period: str):
        '''
        Return the attendance information for the concerned student for the 
        concerned month.
        
        Parameters:
            login (string): the username of the student
            period (string): the month in the format YYYY-MM
        
        Returns:
            info (list): A list contaning logID, type of contact, number of
            contacts, and notes.
        '''
        with connection.cursor() as cursor:
            query = "SELECT logID, contactType, contactCount, note FROM \
                ElecEngwww.tier4log WHERE studentLogin = '" + login + "' AND \
                logPeriod = '" + period + "'"
            info = []
            for row in cursor.execute(query).fetchall():
                info.append([x for x in row])
            cursor.close()
            return (info)

    def isAuthorized(eeID:int):
        '''
        Returns whether the user is a supervisor or not
        
        Parameters:
            eeid (int): The ee ID of the user.
            
        Returns:
            boolean (boolean): True if the user is supervisor False otherwise.
        
        '''
        with connection.cursor() as cursor:
            query = "SELECT eeid, CID, login, Namei, FeeClass, ERegyr FROM \
                eedbo.eepx WHERE (EReg = 'P' OR EReg = 'F') and (PSup1 = " + \
                eeID + " or PSup2 = " + eeID + ")"
            pgrList = cursor.execute(query).fetchall()
            if pgrList:
                return(True)
            else:
                return(False)
                
    def getIncompleteListMail():
        '''
        Sends mail to the supervisors if the log of his/her supervisees is incomplete 
        '''
        with connection.cursor() as cursor:
            supervisorList = {}
            query = "SELECT eeid, CID, login, Namei, ERegyr, Alias, PSup1, \
                PSup2, Formal FROM eedbo.eepx WHERE ALIAS = 1 and (EReg = 'P' OR \
                EReg = 'F') and SUBSTRING(ERegyr, 2, 2) < 6"
            pgrList = cursor.execute(query).fetchall()
            lastYear = datetime.datetime.now().year
            lastMonth = datetime.datetime.now().month
            list = []
            for pgr in pgrList:
                query = "SELECT Finish, EEntry FROM eedbo.eepx WHERE Login = '" \
                    + pgr[2] + "'"
                result = cursor.execute(query).fetchone()
                startYear = int(str(result[1])[:4]) if int(str(result[1])[:4]) > T4Functions.year else T4Functions.year + 1
                startMonth = int(str(result[1])[5:7]) if int(str(result[1])[:4]) > T4Functions.year else 1
                for yr in range(startYear, lastYear + 1):
                    br = False
                    for mnth in range(1, 13):
                        if (yr == lastYear and mnth < lastMonth) or (yr == \
                                startYear and mnth >= startMonth) or (yr < \
                                lastYear and yr > startYear):
                            info = T4Functions.getPeriodInfo(pgr[2], \
                                str(yr) + "-" + "{:02d}".format(mnth))
                            if not info:
                                supervisor = CommonFunctions.getUserByEEID(pgr[6])
                                if supervisor[0] not in supervisorList:
                                    supervisorList[supervisor[0]] = {}
                                    supervisorList[supervisor[0]]["supervisor"] = supervisor
                                    supervisorList[supervisor[0]]["supervisee"] = []
                                supervisorList[supervisor[0]]["supervisee"].append(pgr)
                                br = True
                                break
                    if br:
                        break
            for supervisor in supervisorList:
                message = "Dear " + supervisorList[supervisor]["supervisor"][3] + ",\n \
                    Can you please log in the Tier 4 attendance entry for the \
                    following student(s):\n"
                for pgr in supervisorList[supervisor]["supervisee"]:
                    message += pgr[8] + " - https://apps.ee.ic.ac.uk/t4monitoring/contacts/"+str(pgr[0])+"/0000-00 \n"
                message += "\nThanks,\nAnderson"
                list.append(message)
            return(list)
            
    def getIncompleteList(studentLogin: str):
        '''
        Returns the list of months for which the students attendance log 
        is not entered.
        
        Parameters:
            studentLogin (string): The username of the student aginst 
            which the databse is to be queried.
            
        Returns:
            data (dictionary): The dictionary containing the list of 
            months.
        '''
    
        data = {}
        with connection.cursor() as cursor:
            lastYear = datetime.datetime.now().year
            lastMonth = datetime.datetime.now().month
            query = "SELECT MAX(logPeriod) FROM ElecEngwww.tier4log WHERE \
                studentLogin = '" + studentLogin + "'"
            period = cursor.execute(query).fetchone()[0]
            query = "SELECT Finish, EEntry FROM eedbo.eepx WHERE Login = '" \
                + studentLogin + "'"
            result = cursor.execute(query).fetchone()
            startYear = int(str(result[1])[:4]) if int(str(result[1])[:4]) > 2018 else 2019
            startMonth = int(str(result[1])[5:7]) if int(str(result[1])[:4]) > 2018 else 1
            if period:
                lastYear = int(period[:4])
                lastMonth = int(period[5:7])
            else:
                lastYear = int(str(result[1])[:4])
                lastMonth = int(str(result[1])[5:7])
            list = []
            for currentYear in range(startYear, lastYear + 1):
                for currentMonth in range(1, 13):
                    if (currentYear == lastYear and currentMonth <= lastMonth) or (currentYear == \
                            startYear and currentMonth >= startMonth) or (currentYear < \
                            lastYear and currentYear > startYear):
                        info = T4Functions.getPeriodInfo(studentLogin, \
                            str(currentYear) + "-" + "{:02d}".format(currentMonth))
                        if not info:
                            list.append(str(currentYear) + "-" + "{:02d}".format(currentMonth))
            data["incompleteList"] = list
            cursor.close()
            return (data)
        
            
    def getLatestPeriod(studentLogin: str):
        '''
        Returns the latest month for which the students attendance log is
        not entered.
        
        Parameters:
            studentLogin (string): The username of the student aginst 
            which the databse is to be queried.
            
        Returns:
            data (dictionary): The dictionary containing the latest month 
            and year for which the attendance log is not entered or the 
            current month and year if all the attendance have been entered
            and the start month and year.
        '''
    
        data = {}
        with connection.cursor() as cursor:
            curYear = datetime.datetime.now().year
            curMonth = datetime.datetime.now().month
            query = "SELECT MAX(logPeriod) FROM ElecEngwww.tier4log WHERE \
                studentLogin = '" + studentLogin + "'"
            period = cursor.execute(query).fetchone()[0]
            query = "SELECT Finish, EEntry FROM eedbo.eepx WHERE Login = '" \
                + studentLogin + "'"
            result = cursor.execute(query).fetchone()
            startYearDiff = curYear - int(str(result[1])[:4])
            startMonthDiff = curMonth - int(str(result[1])[5:7])
            if startMonthDiff < 0:
                startMonthDiff = 12 + startMonthDiff
                startYearDiff -= 1
            startMonthDiff += startYearDiff*12
            startYearDiff = 0
            if result[0]:
                data["finished"] = True
                dbMonth = int(result[0][5:7])
                dbYear =  int(result[0][:4])
            else:
                data["finished"] = False
                if period:
                    dbYear = int(period[:4])
                    dbMonth = int(period[5:7])
                    if dbMonth == curMonth and dbYear == curYear:
                        data["latest"] = True
                    elif (dbMonth > curMonth and dbYear == curYear) or (dbYear > curYear):
                        data["latest"] = True
                        dbMonth = curMonth
                        dbYear =  curYear
                    else:
                        data["latest"] = False
                        dbYear = dbYear + (int)(dbMonth / 12)
                        dbMonth = (dbMonth % 12) + 1
                else:
                    dbMonth = int(str(result[1])[5:7])
                    dbYear =  int(str(result[1])[:4])
            data.update({"month":"{:02d}".format(dbMonth), "year":dbYear, \
                "startMonth": startMonthDiff, "startYear": startYearDiff})
            cursor.close()
            return(data)

class T4ReportView(View):
    '''
        Page for the superuser to upload the report template and get it 
        updated with new information
    '''
    def get(self, request):
        data = {}
        data.update(CommonFunctions.getUserLogin(request, T4Functions.suAccess))
        data["pageTitle"] = "T4 Reports"
        data["breadcrumb"] = '<li><a href="/t4monitoring/' + data["urlEncode"]\
            + '"> Home</a></li><li><a>' + data["pageTitle"] + '</a></li>'
        if CommonFunctions.getUserLogin(request, T4Functions.suAccess)["user"] not in T4Functions.suAccess:
            return redirect("/t4monitoring/"+data["urlEncode"])
        template = loader.render_to_string('app/report.html', {"data": data})
        return HttpResponse(template)

    def post(self, request):
        excelFile = request.FILES["template"]
        wb = openpyxl.load_workbook(excelFile)
        worksheet = wb.active
        wbError = openpyxl.Workbook()
        sheet = wbError.active
        sheet.cell(row = 1, column = 1).value = "CID"
        sheet.cell(row = 1, column = 2).value = "Login"
        sheet.cell(row = 1, column = 3).value = "Log Period"
        sheet.cell(row = 1, column = 4).value = "Existing"
        sheet.cell(row = 1, column = 5).value = "Generated Record"
        discrepancies = list()
        
        contacts = {"1":"f2f", "2":"remote", "3":"Submitted", "4":"No contact"}
        s = io.BytesIO()
        zf = zipfile.ZipFile(s, mode="w")
        k = 1
        with connection.cursor() as cursor:
            for i, row in enumerate(worksheet.iter_rows()):
                if i > 0:
                    cid = row[0].value
                    if cid is not None:
                        query = "SELECT login FROM eedbo.eepx WHERE CID = " + str(cid)
                        login = cursor.execute(query).fetchone()
                        if login is not None:
                            login = login[0]
                            if login is not None:
                                j = 14
                                while j < len(worksheet[1]) and \
                                    str(worksheet[1][j].value)[5:7] in T4Functions.monthsRev:
                                    logPeriod = "20" + worksheet[1][j].value[4:] \
                                        + "-" + months[worksheet[1][j].value[:3]]
                                    query = "SELECT contactType, contactCount, note FROM \
                                        ElecEngwww.tier4log WHERE studentLogin = '" \
                                        + login + "' and logPeriod = '" + logPeriod + "'"
                                    result =  cursor.execute(query).fetchall()
                                    if result:
                                        flag = False
                                        format = ""
                                        for rowDB in result:
                                            if format is not "":
                                                flag = True
                                                format += "\n"
                                            format += contacts[str(rowDB[0])]
                                            if rowDB[0] < 3:
                                                format += " = " + ("1" if rowDB[1] \
                                                    < 2 else "2 or more")
                                            elif rowDB[0] == 4 and rowDB[2]:
                                                format += " because " + rowDB[2]
                                        if row[j].value != format and row[j].value \
                                            is not None:
                                            k += 1
                                            discrepancies.append([cid, login, \
                                                logPeriod, row[j].value, format])
                                            sheet.cell(row = k, column = 1).value = cid
                                            sheet.cell(row = k, column = 2).value = login
                                            sheet.cell(row = k, column = 3).value = logPeriod
                                            sheet.cell(row = k, column = 4).value = row[j].value
                                            sheet.cell(row = k, column = 5).value = format
                                        worksheet.cell(row = i+1, column = j+1).value = format
                                        if flag:
                                            worksheet.cell(row = i+1, column = \
                                                j+1).alignment = Alignment(wrapText=True)
                                    j += 1
        cursor.close()
        zf.writestr("templateMod.xlsx", save_virtual_workbook(wb))
        zf.writestr("Discrepancies.xlsx", save_virtual_workbook(wbError))
        zf.close()
        response = HttpResponse(s.getvalue(), content_type="application/x-zip-compressed")
        response['Content-Disposition'] = "attachment; filename=processedOP.zip"
        return response
        
class T4DataUploadView(View):
    '''
        Page to synch the database with old information
    '''
    def get(self, request):
        data = {}
        data.update(CommonFunctions.getUserLogin(request, T4Functions.suAccess))
        data["pageTitle"] = "T4 DB Sync"
        data["breadcrumb"] = '<li><a href="/t4monitoring/' + data["urlEncode"]\
            + '"> Home</a></li><li><a>' + data["pageTitle"] + '</a></li>'
        if CommonFunctions.getUserLogin(request, T4Functions.suAccess)["user"] not in T4Functions.suAccess:
            return redirect("/t4monitoring/"+data["urlEncode"])
        template = loader.render_to_string('app/report.html', {"data": data})
        return HttpResponse(template)

    def post(self, request):
        excelFile = request.FILES["template"]
        wb = openpyxl.load_workbook(excelFile)
        worksheet = wb.active
        excelData = list()
        with connection.cursor() as cursor:
            for i, row in enumerate(worksheet.iter_rows()):
                if i > 0:
                    rowData = list()
                    if row[5].value is not None:
                        contactType = 1 if row[6].value == "Face-to-face" else 2
                        count = 2 if row[7].value == "2 or more" else 1
                        period = str(row[5].value.strftime('%Y-%m'))
                        studentLogin = row[1].value
                        updatedBy = "hgupta"
                        query = "SELECT * FROM ElecEngwww.tier4log WHERE \
                            studentLogin = '" + studentLogin + "' and logPeriod = \
                            '" + period + "' and contactType = " + str(contactType)
                        result = cursor.execute(query).fetchall()
                        if not result:
                            query = "INSERT INTO ElecEngwww.tier4log (studentLogin, \
                                logPeriod, contactType, contactCount, updateBy) VALUES \
                                ('" + studentLogin + "', '" + period + "', " + str(contactType) + ", \
                                " + str(count) + ", '" + updatedBy + "')"
                            cursor.execute(query)
                        else:
                            query = "UPDATE ElecEngwww.tier4log SET contactCount = " + \
                                str(result[0][4] + count) + " WHERE logID = " + str(result[0][0])
                            cursor.execute(query)
        cursor.close()
        return HttpResponse("excelData")
        
   
class T4MasqueradeView(View):
    '''
        Page for superuser to select a view_as user
    '''
    def get(self, request):
        data = {}
        data.update(CommonFunctions.getUserLogin(request, T4Functions.suAccess))
        if CommonFunctions.getUserLogin(request, T4Functions.suAccess)["user"] not in T4Functions.suAccess:
            return redirect("/t4monitoring/" + data["urlEncode"])
        with connection.cursor() as cursor:
            data["pageTitle"] = "Run T4 Monitoring system as other user"
            data["breadcrumb"] = '<li><a href="/' + data["urlEncode"] + '">Home</a>\
                </li><li><a>' + data["pageTitle"] + '</a></li>'
            query = "SELECT * FROM eedbo.GCategories ORDER BY GCategory"
            data["GCategories"] = cursor.execute(query).fetchall()
            query = "SELECT * FROM eedbo.Depts ORDER BY 'dept name'"
            data["depts"] = cursor.execute(query).fetchall()
            data["appDomain"] = "t4monitoring"
            template = loader.render_to_string(
                'common/masquerade.html', {"data": data})
            cursor.close()
        return HttpResponse(template)
    
    def post(self, request):
        return HttpResponse("POST")